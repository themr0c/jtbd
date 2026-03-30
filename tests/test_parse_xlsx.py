import sys
from pathlib import Path
import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / 'scripts'))

from lib.parse_xlsx import _cell, parse_jobs, parse_proposal


def _wb(jobs_rows, proposal_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Phase 3 Jobs'
    for row in jobs_rows:
        ws.append(row)
    if proposal_rows is not None:
        ws2 = wb.create_sheet('Phase 3 JTBD - All-in Proposal')
        for row in proposal_rows:
            ws2.append(row)
    return wb


def test_cell_none():
    assert _cell(None) == ''


def test_cell_float():
    assert _cell(3.0) == '3'


def test_cell_string_stripped():
    assert _cell('  hello  ') == 'hello'


def test_parse_jobs_basic():
    header = ['Job #', 'Job', 'Job Statement', 'Job Category ', 'Job Priority (calculated)']
    row1 = [1, 'Install RHDH', 'When I install...', 'Install', 'high']
    jobs = parse_jobs(_wb([header, row1]))
    assert len(jobs) == 1
    assert jobs[0]['num'] == '1'
    assert jobs[0]['name'] == 'Install RHDH'
    assert jobs[0]['priority'] == 'Major'
    assert jobs[0]['category_label'] == 'install'


def test_parse_jobs_skips_empty_num():
    header = ['Job #', 'Job', 'Job Statement', 'Job Category ', 'Job Priority (calculated)']
    jobs = parse_jobs(_wb([header, [None, 'No number', '', '', '']]))
    assert jobs == []


def test_parse_proposal_returns_empty_when_sheet_missing():
    wb = openpyxl.Workbook()
    wb.active.title = 'Phase 3 Jobs'
    assert parse_proposal(wb) == {}
